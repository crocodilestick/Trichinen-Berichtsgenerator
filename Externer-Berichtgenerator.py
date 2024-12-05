import datetime
import json
import os
import sys
from datetime import datetime

from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import load_workbook

# DEBUGGING
# from pprint import pprint


current_dir = os.path.dirname(os.path.abspath(__file__))
auftrag_nr = ""

class TrichExcel():
    def __init__(self, filename) -> None:
        self.wb = load_workbook(filename)
        self.ws = self.wb.active
        self.rows = self.ws.max_row
        self.first_data_row = 5
        self.header_row = 4
        self.cols = self.ws.max_column
        self.letters = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M")
        self.column_headers = self.get_column_headings() 
        self.data = self.get_data()
        if self.data is not None:
            self.auftrag_nr = self.data[0]['Auftrag']

    def get_column_headings(self) -> list[str]:
        column_headers = []
        for letter in self.letters:
            column_headers.append(self.ws[f"{letter}{self.header_row}"].value)

        return column_headers

    def get_data(self) -> list[dict[str, str]]:
        data = []
        x = self.first_data_row
        while x != (self.rows + 1):
            sample = {}
            y = 1
            while y < (self.cols + 1):
                value = self.ws[f"{self.letters[y - 1]}{x}"].value
                if type(value) == datetime:
                    value = value.strftime('%d.%m.%Y')
                elif value != None:
                    value = str(value)
                sample[self.column_headers[y - 1]] = value
                y += 1

            data.append(sample)
            x += 1

        return data

class PDF(FPDF):
    def footer(self):
        # Go to 1.5 cm from bottom
        self.set_y(-10)
        # Select Helvetica Bold Size=8
        self.set_font('Helvetica', 'B', 8)
        # Print footer image
        self.image(f"{current_dir}/assets/footer.png", x=10, y=192, h=12.5)
        # Print page number
        self.cell(282, None, ((f"Auftrag: {auftrag_nr}       "
                               f"Seite {self.page_no()} "
                                "von {nb}")),
                                align='R')

class ReportMaker():
    def __init__(self, pdf_object, lab_info, excel_data, prufleiter, recipient_address):
        self.pdf = pdf_object
        self.excel_data = excel_data
        self.headers = FontFace(emphasis="BOLD", color=(0,0,0), fill_color=(191,191,191))

        self.lab_info = lab_info

        self.prufleiter_name = prufleiter["name"]
        self.prufleiter_email = prufleiter["email"]
        self.prufleiter_tel = prufleiter["tel"]


        self.recipient_address = recipient_address

        self.column_widths = (1, 1.5, 1.6, 1.5, 1.5, 2, 1.2, 2, 1.4)
        self.columns = ("Labor ID",
                       "Probeneingang im Labor",
                       "Datum Probenentnahme",
                       "Probenehmer",
                       "Tierart",
                       "Wildmarkennummer/ Schlachtnummer",
                       "Parameter",
                       "Ergebnis - Wert",
                       "Probenmenge")

    def add_header(self):
        self.pdf.image(f"{current_dir}/assets/logo-1.jpg", x=15, y=5, h=15)
        self.pdf.image(f"{current_dir}/assets/logo-2.png", x=245, y=5, h=15)
        self.pdf.ln(12)
        self.pdf.set_font("Helvetica", size=6, style="B")
        self.pdf.cell(0, None, self.lab_info['address'], align='L')
        self.pdf.set_font("Helvetica", size=14, style="B")
        self.pdf.cell(-55, None, self.lab_info['department'], align='R')
        self.pdf.set_font("Noto", size=7)
        self.pdf.set_xy(x=202, y=35)
        self.pdf.multi_cell(w=90,
                       align='L',
                       text=(f"{'Datum:': <10} {datetime.today().strftime('%d.%m.%Y'): >45}\n"
                             f"{'Bearb:': <10} {self.prufleiter_name: >45}\n"
                             f"{'Tel:': <10} {self.prufleiter_tel: >45}\n"
                             f"{'Fax:': <10} {self.lab_info['fax_num']: >45}\n"
                             f"{'Email:': <10} {self.prufleiter_email: >45}\n"
                             f"{'Internet:': <10} {self.lab_info['website']: >45}\n\n"
                             f"{'Auftrag:': <10} {self.excel_data.data[0]['Auftrag']: >45}"
                            )
                      )
        self.pdf.set_xy(x=10, y=30)
        self.pdf.set_font("Helvetica", size=14, style="B")
        self.pdf.cell(0, None, self.recipient_address[0], align='L')
        self.pdf.ln(10)
        self.pdf.set_font("Helvetica", size=12)
        self.pdf.cell(0, None, self.recipient_address[1], align='L')
        self.pdf.ln()
        self.pdf.cell(0, None, self.recipient_address[2], align='L')
        self.pdf.ln(16)

    def add_title_table(self):
        self.pdf.set_font("Helvetica", size=9.5)

        with self.pdf.table(text_align="CENTER", col_widths=self.column_widths) as table:
            row = table.row()
            row.cell(("Ergebnisbericht Trichinenuntersuchung / Untersuchung auf Duncker'"
                      "scher Muskelegel (Alaria alata)*"),
                     colspan=len(self.column_widths))

            row = table.row()
            self.pdf.set_font("Helvetica", size=7)
            row.cell(("Methode Trichinenuntersuchung: Magnetrührverfahren für die künst"
                      "liche Verdauung nach DIN EN ISO 18743:2015 (2015-12); Methode Du"
                      "ncker'scher Muskelegel: PV3764 (2015-02; Hausmethode)"),
                     colspan=len(self.column_widths))

            self.pdf.set_font("Helvetica", size=9.5)
            row = table.row()
            row.cell(f"{self.lab_info['kurzel']} Probendaten", colspan=2)
            row.cell("Probendaten vom Einsender zur Verfügung gestellt", colspan=4)
            row.cell("Ergebnisse", colspan=3)

    def add_data_table(self):
        with self.pdf.table(text_align="CENTER", col_widths=self.column_widths) as table:
            row = table.row(style=self.headers)
            for column in self.columns:
                row.cell(column)

            self.pdf.set_font("Helvetica", size=7)
            excluded_columns = ["Auftrag",
                                "Bemerkung",
                                "Ergebnisvalidation Prüfleitung",
                                "Datum Ergebnisvalidation"]

            for sample in self.excel_data.data:
                if sample["Parameter"] == "Trichinella sp." or sample["Parameter"] == "Alaria alata":
                    row = table.row()
                    for header in self.excel_data.column_headers:
                        if header not in excluded_columns:
                            if header == "Probenmenge (in g)/ Tier":
                                row.cell(str(sample[header]) + "g/Tier")
                            else:
                                row.cell(sample[header])

    def add_bemerkungen(self):
        self.pdf.ln()
        self.pdf.set_font("Helvetica", size=14, style="B")
        self.pdf.cell(0, None, "Proben mit Bemerkungen", align='L')
        self.pdf.ln(8)

        self.pdf.set_font("Helvetica", size=7)

        with self.pdf.table(text_align="LEFT", col_widths=(1, 9)) as table:
            row = table.row(style=self.headers)
            row.cell("Labor ID")
            row.cell("Bemerkung")

            for sample in self.excel_data.data:
                if sample["Bemerkung"] is not None:
                    row = table.row()
                    row.cell(sample["Labor ID"])
                    row.cell(sample["Bemerkung"])

    def add_end_of_report(self):
        self.pdf.ln()
        self.pdf.set_font("Helvetica", size=10, style="B")
        self.pdf.cell(0, None, f"Ergebnisbericht vom:     {datetime.today().strftime('%d.%m.%Y')}", align='L')
        self.pdf.ln(6)
        self.pdf.cell(0, None, f"Validiert von:                    {self.prufleiter_name}", align='L')
        self.pdf.ln(6)
        self.pdf.set_font("Helvetica", size=7, style="I")
        self.pdf.cell(0, None, "   Das Dokument wurde elektronisch schlussgezeichnet und ist ohne Unterschrift gültig.", align='L')
        self.pdf.ln(6)
        self.pdf.set_font("Helvetica", size=9, style="I")
        self.pdf.cell(0, None, "Die Prüfergebnisse beziehen sich ausschließlich auf die erhaltenen und untersuchten Proben.", align='L')
        self.pdf.ln()
        self.pdf.cell(0, None, f"Eine auszugsweise Veröffentlichung oder Vervielfältigung des Berichtes ist nur mit schriftlicher Genehmigung des {self.lab_info['name']} erlaubt.", align='L')
        self.pdf.ln(6)
        self.pdf.cell(0, None, "* Die Untersuchung auf DME läuft außerhalb der Akkreditierung", align='L')

    def save(self, filename):
        self.pdf.output(filename)

def main():
    print("-------------------------------------")
    print("|                                   |")
    print("|      Extern Berichtsgenerator     |")
    print("|                                   |")
    print("-------------------------------------")

    try:
        if len(sys.argv) < 2:
            print("\nKeine Datei zur Verarbeitung gegeben.\n".upper())
            print(("Um dieses Programm zu verwenden, ziehen Sie einfach die Excel-Datei mit den\n"
                  "Daten, die Sie aus Limsophy exportiert haben, auf das Symbol des Programms."))
            input("\nDrücken Sie zum Beenden die Eingabetaste.")
            sys.exit(1)
        elif len(sys.argv) > 2:
            print("\nzu viele Dateien zur Verarbeitung gegeben.\n".upper())
            print(("Dieses Programm kann nur eine Datei auf einmal verarbeiten. Versuchen Sie\n"
                  "erneut, nur eine Datei bereitzustellen."))
            input("\nDrücken Sie zum Beenden die Eingabetaste.")
            sys.exit(2)

        lab_info = get_lab_info()

        excel_file = sys.argv[1]
        excel_filename = os.path.basename(excel_file).replace('.xlsx', '')
        excel_data = TrichExcel(excel_file)
        global auftrag_nr
        auftrag_nr = excel_data.auftrag_nr

        recipient_address = get_recipient_address()

        # DEBUGGING
        # pprint(excel_data.data, sort_dicts=False)
        # input("")

        prufleiter = get_prufleiter_data()
        print(f"\nWillkommen {prufleiter['name']}! Bericht generieren...")

        pdf = PDF(orientation="landscape", format="A4")
        pdf.add_font(family="Noto", fname=f"{current_dir}/assets/NotoSansMono-Regular.ttf")
        pdf.add_page()

        report = ReportMaker(pdf, lab_info, excel_data, prufleiter, recipient_address)
        report.add_header()
        report.add_title_table()
        report.add_data_table()

        bemerkungen = []
        for sample in excel_data.data:
            if sample["Bemerkung"] is not None:
                bemerkungen.append(sample["Bemerkung"])
        if len(bemerkungen) > 0:
            report.add_bemerkungen()

        report.add_end_of_report()

        print("\nBericht erfolgreich erstellt!")
        print("\nGeben Sie einen Namen für die Berichtsdatei ein oder drücken Sie\n"
              "einfach die Eingabetaste, um denselben Namen wie die ursprüngliche\n"
              "Excel-Datei zu verwenden.\n")
        report_name = input("Dateiname für den Bericht: ").strip().replace(" ", "_")
        if report_name == "":
            report_name = excel_filename
        report.save(report_name + ".pdf")

        sys.exit(0)
    except Exception as e:
        print(f"The following error occurred: {e}")
        input("Press Enter to exit")

def get_prufleiter_data() -> dict[str, str]:
    while True:
        kurzel = input("\nGeben Sie ihre Kurzel ein: ").strip().upper()
        with open(f"{current_dir}/assets/prufleitung.json", "r", encoding="utf-8") as f:
            prufleiter_data = json.load(f)
            if kurzel in prufleiter_data:
                prufleiter = prufleiter_data[kurzel]
                break
            else:
                print(f"\nDie Kurzel '{kurzel}' ist nicht vorhanden.")
                print("Bitte versuchen Sie es erneut.")

    return prufleiter

def get_recipient_address() -> dict[str, str]:
    while True:
        with open(f"{current_dir}/assets/adressen.json", "r", encoding="utf-8") as f:
            adressen = json.load(f)
            print(f"\nVerfügbare Empfänger: {', '.join(adressen.keys())}")
            kurzel = input("\nGeben Sie die Kurzbezeichnung des Empfängers ein: ").strip().lower()
            if kurzel in adressen:
                adresse = adressen[kurzel]
                break
            else:
                print(f"\nDer Empfänger '{kurzel}' ist nicht vorhanden.")
                print("Bitte versuchen Sie es erneut.")
        
    return adresse

def get_lab_info() -> dict[str, str]:
    with open(f"{current_dir}/assets/lab-info.json", "r", encoding="utf-8") as f:
        lab_info = json.load(f)

    return lab_info

if __name__ == "__main__":
    main()